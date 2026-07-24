"""End-to-end FCA workbook profiling orchestration."""

from __future__ import annotations

import json
import logging
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

from customer_harm.profiling.discovery import discover_workbooks, parse_reporting_period, sha256_file
from customer_harm.profiling.definitions import (
    build_product_group_inventory,
    compare_definitions,
    compare_product_groups,
    compare_taxonomies,
    extract_notes_content,
)
from customer_harm.profiling.headers import (
    HEADER_TERM_PROFILES,
    classify_header_confidence,
    detect_header,
    normalise_text,
)
from customer_harm.profiling.schema import canonical_sheet_name, classify_sheet_type, compare_schemas
from customer_harm.profiling.values import profile_values
from customer_harm.profiling.review import (
    build_review_register,
    render_extraction_readiness,
    render_review_report,
    review_gate_summary,
)

LOGGER = logging.getLogger(__name__)
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
EXPECTED_PERIODS = tuple(f"{year}-H{half}" for year in range(2021, 2026) for half in (1, 2))
METRIC_SHEET_ROLES = {
    "opened", "closed", "percentage_within_3_days",
    "percentage_after_3_days_within_8_weeks", "percentage_upheld",
    "consumer_credit", "context_intermediation", "context_provision",
}


class ProfilingRunError(RuntimeError):
    """Raised after auditable outputs are written for an incomplete profiling run."""


@dataclass(frozen=True)
class ProfileConfig:
    """Runtime paths and bounded header-search settings."""

    input_dir: Path = Path("data/raw/fca/firm_level")
    output_dir: Path = Path("data/interim/profiling")
    metadata_dir: Path = Path("data/metadata")
    header_scan_rows: int = 40
    expected_workbook_count: int = 10
    source_inventory_path: Path = Path("docs/source_inventory.csv")
    expected_periods: tuple[str, ...] = EXPECTED_PERIODS
    review_decisions_path: Path = Path("data/mappings/profiling_review_decisions.csv")
    review_report_path: Path | None = None
    extraction_readiness_path: Path | None = None


def _trim_matrix(rows: Iterable[tuple[Any, ...]]) -> tuple[list[list[Any]], int, int]:
    original = [list(row) for row in rows]
    fully_empty_rows = sum(not any(value is not None and str(value).strip() for value in row) for row in original)
    matrix = list(original)
    while matrix and not any(value is not None and str(value).strip() for value in matrix[-1]):
        matrix.pop()
    if not matrix:
        return [], fully_empty_rows, 0
    width = max(len(row) for row in matrix)
    empty_columns = sum(
        not any(column < len(row) and row[column] is not None and str(row[column]).strip() for row in matrix)
        for column in range(width)
    )
    last_used = max(
        (column for row in matrix for column, value in enumerate(row)
         if value is not None and str(value).strip()), default=-1
    )
    return [row[: last_used + 1] for row in matrix], fully_empty_rows, empty_columns


def _make_column_names(header: list[Any]) -> tuple[list[str], list[str]]:
    names: list[str] = []
    duplicates: list[str] = []
    counts: dict[str, int] = {}
    for position, value in enumerate(header, 1):
        base = str(value).strip() if value is not None and str(value).strip() else f"unnamed_{position}"
        counts[base] = counts.get(base, 0) + 1
        if counts[base] > 1:
            duplicates.append(base)
        names.append(base if counts[base] == 1 else f"{base}.{counts[base] - 1}")
    return names, duplicates


def _as_frame(matrix: list[list[Any]], header_index: int) -> tuple[pd.DataFrame, list[str]]:
    if not matrix:
        return pd.DataFrame(), []
    columns, duplicates = _make_column_names(matrix[header_index])
    width = len(columns)
    rows = [(row + [None] * width)[:width] for row in matrix[header_index + 1:]]
    return pd.DataFrame(rows, columns=columns).dropna(how="all").convert_dtypes(), duplicates


def _classify_format(formats: list[str]) -> str:
    combined = " ".join(formats).casefold()
    if "%" in combined:
        return "percentage"
    if any(token in combined for token in ("£", "$", "€", "[$£")):
        return "currency"
    if any(token in combined for token in ("yy", "dd", "mm")):
        return "date"
    return "number" if any(char in combined for char in ("0", "#")) else "general"


def _serialise(value: Any) -> Any:
    if pd.isna(value):
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value.item() if hasattr(value, "item") else value


def _potential_special_rows(
    matrix: list[list[Any]], start_index: int, canonical_sheet: str
) -> tuple[list[int], list[int]]:
    """Find conservative structural labels; never match arbitrary text in data rows."""
    total_rows: list[int] = []
    note_rows: list[int] = []
    for index, row in enumerate(matrix[start_index:], start_index + 1):
        populated = [normalise_text(value) for value in row if value is not None and str(value).strip()]
        if not populated:
            continue
        first_value = populated[0]
        if canonical_sheet in METRIC_SHEET_ROLES and first_value in {"grand total", "total", "subtotal"}:
            total_rows.append(index)
        if canonical_sheet == "notes" and (
            first_value in {"note", "notes", "source", "definition"}
            or re.match(r"^(note|notes|source|definition):\s+", first_value)
        ):
            note_rows.append(index)
    return total_rows, note_rows


def _worksheet_xml_metadata(path: Path) -> dict[str, dict[str, Any]]:
    """Read value bounds, formulas and merged ranges without trusting formatted dimensions."""
    result: dict[str, dict[str, Any]] = {}
    with zipfile.ZipFile(path) as archive:
        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        relationships_root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        targets = {
            relation.attrib["Id"]: relation.attrib["Target"]
            for relation in relationships_root.findall(f"{{{PKG_REL_NS}}}Relationship")
        }
        for sheet in workbook_root.findall(f".//{{{MAIN_NS}}}sheet"):
            name = sheet.attrib["name"]
            relationship_id = sheet.attrib[f"{{{DOC_REL_NS}}}id"]
            target = targets[relationship_id].lstrip("/")
            xml_path = target if target.startswith("xl/") else f"xl/{target}"
            metadata: dict[str, Any] = {
                "last_value_row": 0,
                "last_value_column": 0,
                "formula_cell_count": 0,
                "formula_without_cached_value_count": 0,
                "merged_ranges": [],
            }
            with archive.open(xml_path) as worksheet_xml:
                for _, node in ElementTree.iterparse(worksheet_xml, events=("end",)):
                    if node.tag == f"{{{MAIN_NS}}}mergeCell":
                        metadata["merged_ranges"].append(node.attrib["ref"])
                    elif node.tag == f"{{{MAIN_NS}}}c":
                        has_formula = node.find(f"{{{MAIN_NS}}}f") is not None
                        has_value = (
                            node.find(f"{{{MAIN_NS}}}v") is not None
                            or node.find(f"{{{MAIN_NS}}}is") is not None
                            or has_formula
                        )
                        if has_formula:
                            metadata["formula_cell_count"] += 1
                            cached_value = node.find(f"{{{MAIN_NS}}}v")
                            if cached_value is None or cached_value.text in (None, ""):
                                metadata["formula_without_cached_value_count"] += 1
                        if has_value and "r" in node.attrib:
                            column_letters, row_number = coordinate_from_string(node.attrib["r"])
                            metadata["last_value_row"] = max(metadata["last_value_row"], row_number)
                            metadata["last_value_column"] = max(
                                metadata["last_value_column"], column_index_from_string(column_letters)
                            )
                    # Keep child <v>/<f>/<is> content until the parent cell is
                    # evaluated; clearing children early makes cached formula
                    # values appear absent. Clearing the cell releases the full
                    # subtree after its metadata has been captured.
                    if node.tag in {f"{{{MAIN_NS}}}c", f"{{{MAIN_NS}}}mergeCell"}:
                        node.clear()
            result[name] = metadata
    return result


def profile_workbooks(config: ProfileConfig) -> dict[str, pd.DataFrame]:
    """Profile every workbook/sheet and persist auditable machine-readable outputs."""
    started_at = datetime.now(timezone.utc).isoformat()
    workbooks = discover_workbooks(config.input_dir)
    if not workbooks:
        raise FileNotFoundError(f"No supported workbooks found under {config.input_dir}")
    parsed_periods = [parse_reporting_period(path.name) for path in workbooks]
    actual_periods = set(parsed_periods)
    expected_periods = set(config.expected_periods)
    duplicate_periods = sorted({period for period in parsed_periods if parsed_periods.count(period) > 1})
    missing_periods = sorted(expected_periods - actual_periods)
    unexpected_periods = sorted(actual_periods - expected_periods)
    if duplicate_periods or missing_periods or unexpected_periods:
        raise ValueError(
            "Invalid reporting-period coverage: "
            f"duplicates={duplicate_periods}, missing={missing_periods}, unexpected={unexpected_periods}"
        )
    config.output_dir.mkdir(parents=True, exist_ok=True)
    config.metadata_dir.mkdir(parents=True, exist_ok=True)
    inventory: list[dict[str, Any]] = []
    sheets: list[dict[str, Any]] = []
    columns: list[dict[str, Any]] = []
    candidates_output: list[dict[str, Any]] = []
    format_profiles: list[dict[str, Any]] = []
    taxonomy_records: list[dict[str, Any]] = []
    definition_records: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    for path in workbooks:
        period = parse_reporting_period(path.name)
        source_id = f"fca_firm_level_{period.casefold().replace('-', '_')}"
        LOGGER.info("event=workbook_started file=%s period=%s", path.name, period)
        status = "success"
        workbook_warning_count = 0
        rows_extracted = 0
        preferred_name = f"fca_firm_level_complaints_{period.casefold().replace('-', '_')}{path.suffix.lower()}"
        if path.name != preferred_name:
            warnings.append({"workbook": path.name, "reporting_period": period, "sheet": None,
                             "warning_code": "NON_STANDARD_FILENAME", "severity": "info",
                             "message": f"Preferred convention is {preferred_name}; raw file was not renamed."})
            workbook_warning_count += 1
        sheet_start = len(sheets)
        column_start = len(columns)
        candidate_start = len(candidates_output)
        format_start = len(format_profiles)
        taxonomy_start = len(taxonomy_records)
        definition_start = len(definition_records)
        warning_start = len(warnings)
        workbook = None
        try:
            xml_metadata = _worksheet_xml_metadata(path)
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            for worksheet in workbook.worksheets:
                role = canonical_sheet_name(worksheet.title)
                metadata = xml_metadata.get(worksheet.title, {})
                last_value_row = int(metadata.get("last_value_row", 0))
                last_value_column = int(metadata.get("last_value_column", 0))
                excel_declared_max_row = worksheet.max_row
                excel_declared_max_column = worksheet.max_column
                value_rows: list[tuple[Any, ...]] = []
                format_observations: list[tuple[int, int, str]] = []
                if last_value_row and last_value_column:
                    for cell_row in worksheet.iter_rows(max_row=last_value_row, max_col=last_value_column):
                        value_rows.append(tuple(cell.value for cell in cell_row))
                        for cell in cell_row:
                            if cell.value is not None and cell.number_format:
                                format_observations.append((cell.row, cell.column, str(cell.number_format)))
                matrix, empty_rows, empty_columns = _trim_matrix(value_rows)
                sheet_type = classify_sheet_type(role)
                if sheet_type == "unclassified":
                    warnings.append({"workbook": path.name, "reporting_period": period, "sheet": worksheet.title,
                                     "warning_code": "UNCLASSIFIED_SHEET", "severity": "warning",
                                     "message": f"No structural profile is configured for canonical role {role}."})
                    workbook_warning_count += 1
                    sheet_type = "metric_wide_table"
                header_terms = HEADER_TERM_PROFILES[sheet_type]
                header_index, candidates = detect_header(
                    matrix, config.header_scan_rows, expected_terms=header_terms
                )
                confidence = classify_header_confidence(candidates)
                frame, duplicates = _as_frame(matrix, header_index)
                if sheet_type == "product_reference":
                    source = {"workbook": path.name, "reporting_period": period,
                              "sheet": worksheet.title}
                    sheet_taxonomy, sheet_definitions = extract_notes_content(
                        matrix, header_index, source
                    )
                    taxonomy_records.extend(sheet_taxonomy)
                    definition_records.extend(sheet_definitions)
                formats_by_column: dict[int, dict[str, int]] = {}
                for cell_row_number, cell_column, cell_format in format_observations:
                    if cell_row_number <= header_index + 1:
                        continue
                    counts = formats_by_column.setdefault(cell_column, {})
                    counts[cell_format] = counts.get(cell_format, 0) + 1
                rows_extracted += len(frame)
                leading_empty = next((i for i, row in enumerate(matrix)
                                      if any(value is not None and str(value).strip() for value in row)), len(matrix))
                totals, notes = _potential_special_rows(matrix, header_index + 1, role)
                merged_ranges = metadata.get("merged_ranges", [])
                raw_headers = matrix[header_index] if matrix else []
                sheets.append({
                    "source_id": source_id, "workbook": path.name, "reporting_period": period,
                    "sheet": worksheet.title, "canonical_sheet": role, "sheet_type": sheet_type,
                    "header_term_profile": sheet_type,
                    "excel_declared_max_row": excel_declared_max_row,
                    "excel_declared_max_column": excel_declared_max_column,
                    "last_value_row": last_value_row, "last_value_column": last_value_column,
                    "used_rows": len(matrix), "used_columns": max((len(row) for row in matrix), default=0),
                    "header_row": header_index + 1, "header_score": candidates[0]["score"] if candidates else None,
                    **confidence,
                    "empty_leading_rows": leading_empty, "fully_empty_rows": empty_rows,
                    "fully_empty_columns": empty_columns, "merged_cell_count": len(merged_ranges),
                    "merged_cells": json.dumps(merged_ranges), "duplicate_column_count": len(duplicates),
                    "duplicate_columns": json.dumps(duplicates), "raw_headers": json.dumps(raw_headers, default=str),
                    "extracted_headers": json.dumps(list(frame.columns), ensure_ascii=False),
                    "data_rows": len(frame), "data_columns": len(frame.columns),
                    "formula_cell_count": int(metadata.get("formula_cell_count", 0)),
                    "formula_without_cached_value_count": int(
                        metadata.get("formula_without_cached_value_count", 0)
                    ),
                    "potential_total_rows": json.dumps(totals), "potential_note_rows": json.dumps(notes),
                })
                for rank, candidate in enumerate(candidates, 1):
                    candidates_output.append({"workbook": path.name, "reporting_period": period,
                                              "sheet": worksheet.title, "sheet_type": sheet_type,
                                              "header_term_profile": sheet_type, "candidate_rank": rank,
                                              "selected": rank == 1, **candidate})
                if not candidates:
                    warnings.append({"workbook": path.name, "reporting_period": period, "sheet": worksheet.title,
                                     "warning_code": "NO_HEADER_CANDIDATE", "severity": "warning",
                                     "message": "No non-empty row was available for header detection."})
                    workbook_warning_count += 1
                if duplicates:
                    warnings.append({"workbook": path.name, "reporting_period": period, "sheet": worksheet.title,
                                     "warning_code": "DUPLICATE_HEADERS", "severity": "info",
                                     "message": f"Duplicate raw headers: {duplicates}"})
                    workbook_warning_count += 1
                if header_index > 0:
                    warnings.append({"workbook": path.name, "reporting_period": period, "sheet": worksheet.title,
                                     "warning_code": "HEADER_NOT_FIRST_ROW", "severity": "info",
                                     "message": f"Selected header row {header_index + 1}; review ranked candidates."})
                    workbook_warning_count += 1
                if confidence["requires_manual_review"]:
                    warnings.append({"workbook": path.name, "reporting_period": period, "sheet": worksheet.title,
                                     "warning_code": "LOW_HEADER_CONFIDENCE", "severity": "warning",
                                     "message": "Header selection has weak anchors or a small/tied score margin."})
                    workbook_warning_count += 1
                if excel_declared_max_row > max(last_value_row * 2, last_value_row + 1000):
                    warnings.append({"workbook": path.name, "reporting_period": period, "sheet": worksheet.title,
                                     "warning_code": "EXCESSIVE_FORMATTED_RANGE", "severity": "warning",
                                     "message": f"Excel declares {excel_declared_max_row} rows but values end at row {last_value_row}."})
                    workbook_warning_count += 1
                formula_without_cache = int(metadata.get("formula_without_cached_value_count", 0))
                if formula_without_cache:
                    warnings.append({"workbook": path.name, "reporting_period": period, "sheet": worksheet.title,
                                     "warning_code": "FORMULA_WITHOUT_CACHED_VALUE", "severity": "warning",
                                     "message": f"{formula_without_cache} formula cells have no cached result."})
                    workbook_warning_count += 1
                for position, column in enumerate(frame.columns, 1):
                    series = frame[column]
                    non_missing = series.dropna()
                    excel_column = get_column_letter(position)
                    format_counts = formats_by_column.get(position, {"General": 0})
                    formats = sorted(format_counts)
                    dominant_format = max(format_counts, key=format_counts.get)
                    format_class = _classify_format(formats)
                    examples = [_serialise(value) for value in non_missing.head(3).tolist()]
                    value_profile = profile_values(series, column, format_class)
                    columns.append({
                        "workbook": path.name, "reporting_period": period, "sheet": worksheet.title,
                        "canonical_sheet": role, "sheet_type": sheet_type,
                        "column_position": position, "excel_column": excel_column,
                        "column": column, "normalised_column": normalise_text(column),
                        "storage_dtype": str(series.dtype), "dtype": str(series.dtype),
                        "row_count": len(series), "non_missing_count": int(series.notna().sum()),
                        "missing_count": int(series.isna().sum()),
                        "missing_percent": round(float(series.isna().mean() * 100), 4) if len(series) else 0.0,
                        "unique_count": int(non_missing.nunique(dropna=True)),
                        "number_formats": json.dumps(formats), "dominant_number_format": dominant_format,
                        "number_format_class": format_class,
                        "example_values": json.dumps(examples, ensure_ascii=False, default=str),
                        **value_profile,
                    })
                    if value_profile["semantic_type"] == "mixed":
                        warnings.append({"workbook": path.name, "reporting_period": period,
                                         "sheet": worksheet.title, "warning_code": "MIXED_VALUE_TYPES",
                                         "severity": "warning",
                                         "message": f"Column {column!r} contains mixed substantive value types."})
                        workbook_warning_count += 1
                    if value_profile["excel_error_count"]:
                        warnings.append({"workbook": path.name, "reporting_period": period,
                                         "sheet": worksheet.title, "warning_code": "EXCEL_ERROR_VALUES",
                                         "severity": "warning",
                                         "message": f"Column {column!r} contains {value_profile['excel_error_count']} Excel errors."})
                        workbook_warning_count += 1
                    marker_count = (
                        value_profile["missing_marker_count"]
                        + value_profile["suppressed_marker_count"]
                        + value_profile["not_applicable_marker_count"]
                    )
                    if marker_count:
                        warnings.append({"workbook": path.name, "reporting_period": period,
                                         "sheet": worksheet.title, "warning_code": "CANDIDATE_MISSING_MARKERS",
                                         "severity": "info",
                                         "message": f"Column {column!r} contains {marker_count} non-null missing/suppression markers."})
                        workbook_warning_count += 1
                    formatted_total = sum(format_counts.values())
                    for number_format, count in sorted(format_counts.items()):
                        format_profiles.append({
                            "workbook": path.name, "reporting_period": period,
                            "sheet": worksheet.title, "canonical_sheet": role,
                            "sheet_type": sheet_type, "column_position": position,
                            "column": column, "number_format": number_format,
                            "number_format_class": _classify_format([number_format]),
                            "formatted_cell_count": count,
                            "formatted_cell_percent": round(count / formatted_total * 100, 4)
                            if formatted_total else 0.0,
                        })
        except Exception as exc:
            status = "failed"
            workbook_warning_count += 1
            del sheets[sheet_start:]
            del columns[column_start:]
            del candidates_output[candidate_start:]
            del format_profiles[format_start:]
            del taxonomy_records[taxonomy_start:]
            del definition_records[definition_start:]
            del warnings[warning_start:]
            rows_extracted = 0
            warnings.append({"workbook": path.name, "reporting_period": period, "sheet": None,
                             "warning_code": "WORKBOOK_FAILURE", "severity": "error", "message": str(exc)})
            LOGGER.exception("event=workbook_failed file=%s", path.name)
        finally:
            if workbook is not None:
                workbook.close()
        inventory.append({
            "source_id": source_id, "source_name": "FCA firm-level complaints", "source_file_name": path.name,
            "reporting_period": period, "file_path": path.relative_to(config.input_dir.parent.parent.parent).as_posix(),
            "file_format": path.suffix.lower().lstrip("."), "file_size_bytes": path.stat().st_size,
            "sha256_hash": sha256_file(path), "ingestion_timestamp": started_at,
            "sheet_count": len(sheet_names) if status == "success" else 0,
            "sheet_names": json.dumps(sheet_names, ensure_ascii=False) if status == "success" else "[]",
            "processing_status": status, "rows_extracted": rows_extracted, "rows_rejected": 0,
            "warning_count": workbook_warning_count, "source_version": period, "revision_note": "",
        })

    frames = {
        "workbook_inventory": pd.DataFrame(inventory),
        "sheet_inventory": pd.DataFrame(sheets),
        "column_profiles": pd.DataFrame(columns),
        "header_detection_results": pd.DataFrame(candidates_output),
        "cell_format_profiles": pd.DataFrame(format_profiles),
        "product_taxonomy_inventory": pd.DataFrame(taxonomy_records, columns=[
            "workbook", "reporting_period", "sheet", "source_row_number",
            "raw_product_group", "inherited_product_group", "raw_product_service_name",
            "normalised_product_service_name", "footnote_marker", "group_was_inherited",
        ]),
        "reporting_definitions": pd.DataFrame(definition_records, columns=[
            "workbook", "reporting_period", "sheet", "source_row_number", "definition_key",
            "definition_type", "section", "definition_text", "definition_text_hash",
        ]),
        "profiling_warnings": pd.DataFrame(warnings, columns=["workbook", "reporting_period", "sheet",
                                                                    "warning_code", "severity", "message"]),
    }
    frames["product_group_inventory"] = build_product_group_inventory(frames["column_profiles"])
    notes_periods = set(frames["sheet_inventory"].loc[
        frames["sheet_inventory"].canonical_sheet == "notes", "reporting_period"
    ])
    frames["product_group_comparison"] = compare_product_groups(frames["product_group_inventory"])
    frames["product_taxonomy_comparison"] = compare_taxonomies(
        frames["product_taxonomy_inventory"], notes_periods, config.expected_periods
    )
    frames["reporting_definition_comparison"] = compare_definitions(
        frames["reporting_definitions"], notes_periods, config.expected_periods
    )
    frames["schema_comparison"] = compare_schemas(frames["column_profiles"], frames["sheet_inventory"])
    frames["profiling_review_register"] = build_review_register(frames, config.review_decisions_path)
    gate = review_gate_summary(frames["profiling_review_register"])
    for name, frame in frames.items():
        frame.to_csv(config.output_dir / f"{name}.csv", index=False)
    review_report = render_review_report(frames["profiling_review_register"], gate)
    readiness_report = render_extraction_readiness(frames["profiling_review_register"], gate)
    (config.output_dir / "profiling_review_report.md").write_text(review_report, encoding="utf-8")
    (config.output_dir / "extraction_readiness.md").write_text(readiness_report, encoding="utf-8")
    (config.output_dir / "profiling_review_gate.json").write_text(
        json.dumps(gate, indent=2) + "\n", encoding="utf-8"
    )
    for report_path, content in (
        (config.review_report_path, review_report),
        (config.extraction_readiness_path, readiness_report),
    ):
        if report_path is not None:
            report_path.parent.mkdir(parents=True, exist_ok=True)
            report_path.write_text(content, encoding="utf-8")
    frames["workbook_inventory"].to_csv(config.metadata_dir / "ingestion_manifest.csv", index=False)
    config.source_inventory_path.parent.mkdir(parents=True, exist_ok=True)
    frames["workbook_inventory"].to_csv(config.source_inventory_path, index=False)
    summary = {
        "profiling_timestamp": started_at, "workbooks_discovered": len(workbooks),
        "expected_workbooks": config.expected_workbook_count,
        "expected_count_met": len(workbooks) == config.expected_workbook_count,
        "expected_periods": sorted(expected_periods),
        "missing_periods": missing_periods, "unexpected_periods": unexpected_periods,
        "duplicate_periods": duplicate_periods,
        "periods": sorted(item["reporting_period"] for item in inventory),
        "sheets_profiled": len(sheets), "columns_profiled": len(columns),
        "warnings": len(warnings), "schema_changes": len(frames["schema_comparison"]),
        "product_group_records": len(frames["product_group_inventory"]),
        "product_group_changes": len(frames["product_group_comparison"]),
        "product_taxonomy_records": len(frames["product_taxonomy_inventory"]),
        "product_taxonomy_changes": len(frames["product_taxonomy_comparison"]),
        "reporting_definition_records": len(frames["reporting_definitions"]),
        "reporting_definition_changes": len(frames["reporting_definition_comparison"]),
        "profiling_review_status": gate["profiling_review_status"],
        "extraction_authorised": gate["extraction_authorised"],
        "blocking_review_items": gate["blocking_review_items"],
        "pending_blocking_items": gate["pending_blocking_items"],
        "failed_workbooks": sum(item["processing_status"] == "failed" for item in inventory),
    }
    (config.output_dir / "profiling_summary.json").write_text(
        json.dumps(summary, indent=2) + "\n", encoding="utf-8"
    )
    LOGGER.info("event=profiling_complete workbooks=%d sheets=%d output=%s", len(workbooks), len(sheets), config.output_dir)
    if any(item["processing_status"] == "failed" for item in inventory):
        raise ProfilingRunError("One or more workbooks failed; partial workbook results were excluded.")
    return frames
