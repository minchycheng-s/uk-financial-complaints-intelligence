"""Build decision-safe BOE, ONS and FOS external-context tables."""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

FCA_PRODUCT_GROUPS = {
    "banking_and_credit_cards",
    "consumer_credit",
    "decumulation_and_pensions",
    "home_finance",
    "insurance_and_pure_protection",
    "investments",
}


def reporting_period(value: pd.Timestamp) -> str:
    """Return the FCA-style half-year containing a timestamp."""
    return f"{value.year}-H{'1' if value.month <= 6 else '2'}"


def period_bounds(period: str) -> tuple[pd.Timestamp, pd.Timestamp]:
    match = re.fullmatch(r"(\d{4})-H([12])", str(period))
    if not match:
        raise ValueError(f"Invalid reporting period: {period}")
    year, half = int(match.group(1)), int(match.group(2))
    if half == 1:
        return pd.Timestamp(year, 1, 1), pd.Timestamp(year, 6, 30)
    return pd.Timestamp(year, 7, 1), pd.Timestamp(year, 12, 31)


def periods_between(start: str, end: str) -> list[str]:
    start_match = re.fullmatch(r"(\d{4})-H([12])", start)
    end_match = re.fullmatch(r"(\d{4})-H([12])", end)
    if not start_match or not end_match:
        raise ValueError("External-context reporting-period bounds are invalid.")
    year, half = int(start_match.group(1)), int(start_match.group(2))
    end_key = (int(end_match.group(1)), int(end_match.group(2)))
    values = []
    while (year, half) <= end_key:
        values.append(f"{year}-H{half}")
        year, half = (year + 1, 1) if half == 2 else (year, 2)
    return values


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build_source_profiles(boe_path: Path, ons_path: Path, fos_path: Path) -> pd.DataFrame:
    """Inventory external sources without modifying them."""
    records: list[dict[str, Any]] = []
    for source, path in [("boe", boe_path), ("ons", ons_path)]:
        frame = pd.read_csv(path, low_memory=False)
        records.append({
            "source": source, "path": path.as_posix(), "file_name": path.name,
            "file_type": path.suffix.lower(), "sha256": sha256(path),
            "sheet_name": "", "rows": len(frame), "columns": len(frame.columns),
        })
    workbook = load_workbook(fos_path, read_only=True, data_only=True)
    fos_hash = sha256(fos_path)
    for sheet in workbook.worksheets:
        records.append({
            "source": "fos", "path": fos_path.as_posix(), "file_name": fos_path.name,
            "file_type": fos_path.suffix.lower(), "sha256": fos_hash,
            "sheet_name": sheet.title, "rows": sheet.max_row, "columns": sheet.max_column,
        })
    return pd.DataFrame(records)


def build_boe_half_year_context(boe: pd.DataFrame, periods: list[str]) -> pd.DataFrame:
    """Summarise effective Bank Rate over each FCA half-year."""
    required = {"Date Changed", "Rate"}
    if not required.issubset(boe.columns):
        raise ValueError(f"BOE source is missing columns: {sorted(required - set(boe.columns))}")
    changes = boe.rename(columns={"Date Changed": "effective_date", "Rate": "bank_rate"}).copy()
    changes["effective_date"] = pd.to_datetime(
        changes["effective_date"], format="%d %b %y", errors="coerce"
    )
    changes["bank_rate"] = pd.to_numeric(changes["bank_rate"], errors="coerce")
    changes = changes.dropna(subset=["effective_date", "bank_rate"]).sort_values("effective_date")
    if changes.empty:
        raise ValueError("BOE source contains no valid rate observations.")

    records = []
    for period in periods:
        start, end = period_bounds(period)
        before_start = changes.loc[changes.effective_date.le(start)]
        before_end = changes.loc[changes.effective_date.le(end)]
        if before_start.empty or before_end.empty:
            raise ValueError(f"BOE history does not cover {period}.")
        start_rate = float(before_start.iloc[-1].bank_rate)
        end_rate = float(before_end.iloc[-1].bank_rate)
        in_period = changes.loc[changes.effective_date.between(start, end)]
        path_values = pd.concat([
            pd.Series([start_rate], dtype=float), in_period.bank_rate.reset_index(drop=True)
        ])
        records.append({
            "reporting_period": period,
            "period_start": start.date().isoformat(),
            "period_end": end.date().isoformat(),
            "start_bank_rate": start_rate,
            "end_bank_rate": end_rate,
            "minimum_bank_rate": float(path_values.min()),
            "maximum_bank_rate": float(path_values.max()),
            "bank_rate_change": end_rate - start_rate,
            "rate_decision_count": int(len(in_period)),
            "measurement_unit": "percentage_points",
            "source_name": "Bank of England",
        })
    return pd.DataFrame(records)


def build_ons_context(
    ons: pd.DataFrame,
    series_config: dict[str, dict[str, str]],
    periods: list[str],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Select documented ONS series and aggregate monthly values to half-years."""
    if "Title" not in ons:
        raise ValueError("ONS source is missing its Title date column.")
    missing = [
        details["source_column"] for details in series_config.values()
        if details["source_column"] not in ons.columns
    ]
    if missing:
        raise ValueError(f"Configured ONS columns are missing: {missing}")
    dates = pd.to_datetime(ons["Title"], format="%Y %b", errors="coerce")
    selected_periods = set(periods)
    records = []
    for indicator, details in series_config.items():
        values = pd.to_numeric(ons[details["source_column"]], errors="coerce")
        for date, value in zip(dates, values):
            if pd.isna(date) or pd.isna(value):
                continue
            period = reporting_period(date)
            if period not in selected_periods:
                continue
            records.append({
                "observation_date": date,
                "reporting_period": period,
                "indicator_name": indicator,
                "indicator_value": float(value),
                "measurement_unit": details["unit"],
                "decision_use": details["decision_use"],
                "source_series": details["source_column"],
                "source_name": "Office for National Statistics",
            })
    monthly = pd.DataFrame(records)
    if monthly.empty:
        raise ValueError("ONS selection produced no monthly context observations.")
    monthly = monthly.sort_values(["indicator_name", "observation_date"]).reset_index(drop=True)

    half_year_records = []
    for (period, indicator), group in monthly.groupby(
        ["reporting_period", "indicator_name"], sort=True
    ):
        group = group.sort_values("observation_date")
        first, last = group.iloc[0], group.iloc[-1]
        half_year_records.append({
            "reporting_period": period,
            "indicator_name": indicator,
            "period_mean": float(group.indicator_value.mean()),
            "period_end_value": float(last.indicator_value),
            "period_minimum": float(group.indicator_value.min()),
            "period_maximum": float(group.indicator_value.max()),
            "within_period_change": float(last.indicator_value - first.indicator_value),
            "month_count": int(len(group)),
            "measurement_unit": last.measurement_unit,
            "decision_use": last.decision_use,
            "source_series": last.source_series,
            "source_name": last.source_name,
        })
    half_year = pd.DataFrame(half_year_records)
    numeric_columns = [
        "period_mean", "period_end_value", "period_minimum",
        "period_maximum", "within_period_change",
    ]
    half_year[numeric_columns] = half_year[numeric_columns].round(6)
    monthly["observation_date"] = monthly.observation_date.dt.date.astype(str)
    return monthly, half_year


def _normalise_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def build_fos_taxonomy(fos_path: Path) -> pd.DataFrame:
    """Convert the FOS old/new taxonomy workbook into a tidy table."""
    workbook = load_workbook(fos_path, read_only=True, data_only=True)
    records = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(min_row=3, values_only=True))
        layouts = [("old", 1, 2, 3), ("new", 5, 6, 7)]
        if sheet.max_column < 8:
            layouts = [("new", 1, 2, 3)]
        for taxonomy_version, sector_i, group_i, type_i in layouts:
            previous_sector, previous_group = "", ""
            for row_number, row in enumerate(rows, start=3):
                sector = _normalise_text(row[sector_i] if len(row) > sector_i else "")
                product_group = _normalise_text(row[group_i] if len(row) > group_i else "")
                product_type = _normalise_text(row[type_i] if len(row) > type_i else "")
                previous_sector = sector or previous_sector
                previous_group = product_group or previous_group
                sector, product_group = sector or previous_sector, product_group or previous_group
                if not any([sector, product_group, product_type]):
                    continue
                records.append({
                    "fos_sheet": sheet.title,
                    "taxonomy_version": taxonomy_version,
                    "fos_sector": sector,
                    "fos_product_group": product_group,
                    "fos_product_type": product_type,
                    "source_row": row_number,
                })
    return pd.DataFrame(records).drop_duplicates().reset_index(drop=True)


def suggest_fca_product_group(row: pd.Series) -> tuple[str, str]:
    """Return a conservative taxonomy suggestion and its reason."""
    text = " ".join([
        str(row.fos_sheet), str(row.fos_sector),
        str(row.fos_product_group), str(row.fos_product_type),
    ]).casefold()
    if any(term in text for term in ["mortgage", "home finance"]):
        return "home_finance", "mortgage_or_home_finance_keyword"
    if any(term in text for term in ["consumer credit", "hire purchase", "payday", "credit broking"]):
        return "consumer_credit", "consumer_credit_keyword"
    if any(term in text for term in ["pension", "annuit", "drawdown"]):
        return "decumulation_and_pensions", "pension_or_annuity_keyword"
    if any(term in text for term in ["investment", "stock", "fund", "bond"]):
        return "investments", "investment_keyword"
    if any(term in text for term in ["insurance", "funeral plan", "protection"]):
        return "insurance_and_pure_protection", "insurance_or_protection_keyword"
    if any(term in text for term in ["banking", "payment", "current account", "savings"]):
        return "banking_and_credit_cards", "banking_or_payment_keyword"
    return "", "no_safe_suggestion"


def build_fca_fos_mapping_template(taxonomy: pd.DataFrame) -> pd.DataFrame:
    """Create mapping candidates that require explicit human review."""
    rows = []
    for record in taxonomy.loc[taxonomy.taxonomy_version.eq("new")].to_dict("records"):
        suggestion, basis = suggest_fca_product_group(pd.Series(record))
        key = "|".join([
            record["fos_sheet"], record["fos_sector"],
            record["fos_product_group"], record["fos_product_type"],
        ])
        rows.append({
            "mapping_id": "FOS-" + hashlib.sha1(key.encode("utf-8")).hexdigest()[:10].upper(),
            "fos_sheet": record["fos_sheet"],
            "fos_sector": record["fos_sector"],
            "fos_product_group": record["fos_product_group"],
            "fos_product_type": record["fos_product_type"],
            "suggested_fca_product_group": suggestion,
            "suggestion_basis": basis,
            "review_decision": "pending_review",
            "approved_fca_product_group": "",
            "reviewer": "",
            "reviewed_at": "",
            "review_comment": "",
        })
    return pd.DataFrame(rows).drop_duplicates("mapping_id").sort_values(
        ["fos_sheet", "fos_sector", "fos_product_group", "fos_product_type"]
    )


def validate_outputs(
    boe: pd.DataFrame,
    ons_monthly: pd.DataFrame,
    ons_half_year: pd.DataFrame,
    taxonomy: pd.DataFrame,
    mappings: pd.DataFrame,
    periods: list[str],
    indicator_count: int,
) -> list[dict[str, str]]:
    checks = {
        "boe_periods_complete": set(boe.reporting_period) == set(periods),
        "boe_rate_ranges_valid": boe[
            ["start_bank_rate", "end_bank_rate", "minimum_bank_rate", "maximum_bank_rate"]
        ].ge(0).all().all(),
        "ons_monthly_grain_unique": not ons_monthly.duplicated(
            ["observation_date", "indicator_name"]
        ).any(),
        "ons_half_year_grain_unique": not ons_half_year.duplicated(
            ["reporting_period", "indicator_name"]
        ).any(),
        "ons_period_indicator_coverage": len(ons_half_year) == len(periods) * indicator_count,
        "fos_taxonomy_not_empty": not taxonomy.empty,
        "fos_mapping_ids_unique": mappings.mapping_id.is_unique,
        "fos_suggestions_valid": set(
            mappings.suggested_fca_product_group.dropna()
        ).difference({""}).issubset(FCA_PRODUCT_GROUPS),
        "fos_suggestions_not_approved": mappings.review_decision.eq("pending_review").all(),
    }
    return [
        {"severity": "error", "rule": rule, "message": f"External-context validation failed: {rule}"}
        for rule, passed in checks.items() if not passed
    ]


def run_external_context_build(
    config_path: Path,
    boe_path: Path,
    ons_path: Path,
    fos_path: Path,
    interim_dir: Path,
    output_dir: Path,
    mapping_path: Path,
) -> dict[str, Any]:
    """Run external-context integration and write generated outputs."""
    required = [config_path, boe_path, ons_path, fos_path]
    missing = [path.as_posix() for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Required external-context inputs are missing: {missing}")
    config = json.loads(config_path.read_text(encoding="utf-8"))
    periods = periods_between(
        config["reporting_period_start"], config["reporting_period_end"]
    )
    profiles = build_source_profiles(boe_path, ons_path, fos_path)
    boe = build_boe_half_year_context(pd.read_csv(boe_path), periods)
    ons_monthly, ons_half_year = build_ons_context(
        pd.read_csv(ons_path, low_memory=False), config["ons_series"], periods
    )
    taxonomy = build_fos_taxonomy(fos_path)
    mappings = build_fca_fos_mapping_template(taxonomy)
    issues = validate_outputs(
        boe, ons_monthly, ons_half_year, taxonomy, mappings,
        periods, len(config["ons_series"]),
    )

    interim_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    mapping_path.parent.mkdir(parents=True, exist_ok=True)
    profiles.to_csv(interim_dir / "external_source_profiles.csv", index=False)
    taxonomy.to_csv(interim_dir / "fos_product_taxonomy.csv", index=False)
    boe.to_csv(output_dir / "boe_half_year_context.csv", index=False)
    ons_monthly.to_csv(output_dir / "ons_monthly_context.csv", index=False)
    ons_half_year.to_csv(output_dir / "ons_half_year_context.csv", index=False)
    pd.DataFrame(issues, columns=["severity", "rule", "message"]).to_csv(
        output_dir / "external_context_validation.csv", index=False
    )
    mapping_created = False
    if not mapping_path.exists():
        mappings.to_csv(mapping_path, index=False)
        mapping_created = True

    summary = {
        "reporting_periods": periods,
        "source_profile_rows": len(profiles),
        "boe_half_year_rows": len(boe),
        "ons_indicators": list(config["ons_series"]),
        "ons_monthly_rows": len(ons_monthly),
        "ons_half_year_rows": len(ons_half_year),
        "fos_taxonomy_rows": len(taxonomy),
        "fos_mapping_candidates": len(mappings),
        "fos_mapping_template_created": mapping_created,
        "warning_score_integration": False,
        "validation_status": "failed" if issues else "passed",
        "validation_issues": issues,
    }
    (output_dir / "external_context_summary.json").write_text(
        json.dumps(summary, indent=2) + "\n", encoding="utf-8"
    )
    if issues:
        raise ValueError(f"External-context validation failed: {issues}")
    return summary
