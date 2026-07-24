from pathlib import Path

import pytest
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from customer_harm.profiling.pipeline import ProfileConfig, ProfilingRunError, profile_workbooks


def _workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Opened"
    sheet.append(["Publication title"])
    sheet.merge_cells("A1:C1")
    sheet.append([None])
    sheet.append(["Firm name", "Product", "Complaints opened"])
    sheet.append(["Bank A", "Credit cards", 10])
    sheet.append(["Bank B", "Mortgages", None])
    workbook.save(path)


def test_pipeline_writes_auditable_outputs(tmp_path: Path) -> None:
    raw, output, metadata = tmp_path / "raw", tmp_path / "profiles", tmp_path / "metadata"
    raw.mkdir()
    _workbook(raw / "fca_2025_h1.xlsx")
    _workbook(raw / "fca_2025_h2.xlsx")

    results = profile_workbooks(ProfileConfig(
        raw, output, metadata, expected_workbook_count=2,
        source_inventory_path=tmp_path / "source_inventory.csv",
        expected_periods=("2025-H1", "2025-H2"),
    ))

    assert results["workbook_inventory"]["sha256_hash"].str.len().eq(64).all()
    assert results["sheet_inventory"]["header_row"].eq(3).all()
    assert results["sheet_inventory"]["merged_cell_count"].eq(1).all()
    assert results["sheet_inventory"]["sheet_type"].eq("metric_wide_table").all()
    assert "semantic_type" in results["column_profiles"].columns
    assert "null_count" in results["column_profiles"].columns
    assert not results["cell_format_profiles"].empty
    assert (output / "schema_comparison.csv").is_file()
    assert (output / "cell_format_profiles.csv").is_file()
    assert (output / "product_group_inventory.csv").is_file()
    assert (output / "product_taxonomy_comparison.csv").is_file()
    assert (output / "reporting_definition_comparison.csv").is_file()
    assert (output / "profiling_review_register.csv").is_file()
    assert (output / "profiling_review_gate.json").is_file()
    assert (output / "profiling_review_report.md").is_file()
    assert (output / "extraction_readiness.md").is_file()
    assert (metadata / "ingestion_manifest.csv").is_file()


def test_formatted_empty_rows_do_not_expand_value_range(tmp_path: Path) -> None:
    raw, output, metadata = tmp_path / "raw", tmp_path / "profiles", tmp_path / "metadata"
    raw.mkdir()
    path = raw / "fca_2025_h1.xlsx"
    _workbook(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Opened"
    sheet.append(["Firm name", "Reporting period", "Banking and credit cards"])
    sheet.append(["Bank A", "2025-H1", 10])
    sheet["A10000"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    workbook.save(path)

    result = profile_workbooks(ProfileConfig(
        raw, output, metadata, expected_workbook_count=1,
        source_inventory_path=tmp_path / "source_inventory.csv",
        expected_periods=("2025-H1",),
    ))

    sheet_profile = result["sheet_inventory"].iloc[0]
    assert sheet_profile["excel_declared_max_row"] == 10000
    assert sheet_profile["last_value_row"] == 2
    assert sheet_profile["fully_empty_rows"] == 0


def test_period_coverage_requires_exact_expected_set(tmp_path: Path) -> None:
    raw = tmp_path / "raw"
    raw.mkdir()
    _workbook(raw / "fca_2025_h1.xlsx")

    with pytest.raises(ValueError, match=r"missing=\['2025-H2'\]"):
        profile_workbooks(ProfileConfig(
            raw, tmp_path / "profiles", tmp_path / "metadata",
            expected_periods=("2025-H1", "2025-H2"),
        ))


def test_failed_workbook_excludes_partial_results_and_raises(tmp_path: Path) -> None:
    raw, output, metadata = tmp_path / "raw", tmp_path / "profiles", tmp_path / "metadata"
    raw.mkdir()
    _workbook(raw / "fca_2025_h1.xlsx")
    (raw / "fca_2025_h2.xlsx").write_bytes(b"not an Excel workbook")

    with pytest.raises(ProfilingRunError):
        profile_workbooks(ProfileConfig(
            raw, output, metadata, expected_workbook_count=2,
            source_inventory_path=tmp_path / "source_inventory.csv",
            expected_periods=("2025-H1", "2025-H2"),
        ))

    inventory = pd.read_csv(output / "workbook_inventory.csv")
    sheets = pd.read_csv(output / "sheet_inventory.csv")
    assert inventory.set_index("reporting_period").loc["2025-H2", "processing_status"] == "failed"
    assert set(sheets["reporting_period"]) == {"2025-H1"}


def test_formula_metadata_distinguishes_missing_cached_value(tmp_path: Path) -> None:
    raw, output, metadata = tmp_path / "raw", tmp_path / "profiles", tmp_path / "metadata"
    raw.mkdir()
    path = raw / "fca_2025_h1.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Opened"
    sheet.append(["Firm name", "Reporting period", "Banking and credit cards"])
    sheet.append(["Bank A", "2025-H1", "=1+1"])
    workbook.save(path)

    result = profile_workbooks(ProfileConfig(
        raw, output, metadata, expected_workbook_count=1,
        source_inventory_path=tmp_path / "source_inventory.csv",
        expected_periods=("2025-H1",),
    ))

    profile = result["sheet_inventory"].iloc[0]
    assert profile["formula_cell_count"] == 1
    assert profile["formula_without_cached_value_count"] == 1
